import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import timedelta
from enviarEmail import enviarEmail
from dropbox_upload import enviar_para_dropbox
import pywhatkit as kit

# Função para obter as informações do usuário
def obter_dados_usuario():
    nome = input("Digite seu nome completo: ").strip()
    while not nome:
        nome = input("Nome inválido. Digite seu nome completo novamente: ").strip()
    
    mes = input("Digite o mês (ex: Janeiro): ").strip()
    while not mes:
        mes = input("Mês inválido. Digite o mês novamente: ").strip()
    
    return nome, mes

# Função para pedir dados das horas extras
def obter_dados_horas(linha):
    # Perguntar ao usuário os dados para a linha atual
    data = input(f"Digite a data (ex: dd/mm/aaaa): ").strip()
    while not validar_data(data):
        data = input(f"Data inválida! Digite novamente (ex: dd/mm/aaaa): ").strip()

    hora_ini = obter_entrada_hora("hora inicial")
    hora_fin = obter_entrada_hora("hora final")
    while hora_fin <= hora_ini:
        print("A hora final deve ser maior que a hora inicial.")
        hora_fin = obter_entrada_hora("hora final")
    
    multiplicador = obter_entrada_multiplicador()
    
    motivo = input(f"Digite o motivo: ").strip()

    # Calcular a diferença de horas (convertendo para timedelta)
    delta_horas = hora_fin - hora_ini
    delta_horas_formatado = format_hora(delta_horas)
    
    return data, hora_ini, hora_fin, multiplicador, delta_horas_formatado, motivo

# Função para validar a data (exemplo simples para dd/mm/aaaa)
def validar_data(data):
    try:
        dia, mes, ano = map(int, data.split('/'))
        return 1 <= dia <= 31 and 1 <= mes <= 12 and ano > 0
    except ValueError:
        return False

# Função para obter a entrada de hora (convertida para float)
def obter_entrada_hora(tipo):
    while True:
        try:
            hora = float(input(f"Digite a {tipo} (em horas decimais, ex: 8.5): ").strip())
            if hora < 0 or hora > 24:
                raise ValueError
            return hora
        except ValueError:
            print(f"Entrada inválida para {tipo}. Digite um valor numérico entre 0 e 24.")

# Função para obter a entrada do multiplicador
def obter_entrada_multiplicador():
    while True:
        try:
            multiplicador = float(input(f"Digite o multiplicador (ex: 1.5): ").strip())
            if multiplicador <= 0:
                raise ValueError
            return multiplicador
        except ValueError:
            print("Entrada inválida para multiplicador. Digite um número positivo.")

# Função para formatar as horas no formato HH:MM:SS
def format_hora(hora_decimal):
    horas = int(hora_decimal)
    minutos = (hora_decimal - horas) * 60
    minutos = round(minutos)
    delta = timedelta(hours=horas, minutes=minutos)
    return str(delta)

# Função para formatar um número decimal em hora (ex: 8.5 para 08:30:00)
def decimal_para_hora(decimal):
    horas = int(decimal)
    minutos = (decimal - horas) * 60
    minutos = round(minutos)
    return f"{horas:02d}:{minutos:02d}:00"

# Criação de uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Horas Extras"

# Definindo o estilo de borda
border = Border(left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"))

# Definindo o estilo de fonte (Calibri, tamanho 14, negrito)
bold_font = Font(name="Calibri", size=14, bold=True)

# Função para definir o texto em caixa alta e aplicar bordas
def set_uppercase(cell, text, alignment='center'):
    cell.value = text.upper()
    cell.font = bold_font
    cell.alignment = Alignment(horizontal=alignment, vertical='center')
    cell.border = border

# Obtendo o nome e o mês do usuário
nome_usuario, mes_usuario = obter_dados_usuario()

# Mesclar e adicionar textos
ws.merge_cells('A1:F1')
set_uppercase(ws['A1'], "PLANILHA DE HORAS EXTRAS")

# Adicionando "NOME" e "MÊS:" nas células A2 e A3 (Centralizado à esquerda)
set_uppercase(ws['A2'], "NOME:", alignment='left')
set_uppercase(ws['A3'], "MÊS:", alignment='left')

# Mesclando e adicionando o nome do usuário em vez de "JOÃO GABRIEL"
ws.merge_cells('B2:E2')
set_uppercase(ws['B2'], nome_usuario, alignment='left')

# Mesclando e adicionando o mês informado em vez de "DEZEMBRO"
ws.merge_cells('B3:E3')
set_uppercase(ws['B3'], mes_usuario, alignment='left')

# Assinatura
ws.merge_cells('D22:E22')
set_uppercase(ws['D22'], "AUTORIZADO POR:", alignment='center')

# Adicionando os cabeçalhos de colunas na linha 4 (Centralizado)
headers = ["DATA", "HORA INI", "HORA FIN", "MULT.", "DELTA(h)", "DESCRIÇÃO"]
for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col_num)
    set_uppercase(cell, header)

# Preenchendo de A5 até F20 com células em branco e aplicando bordas
for row in range(5, 21):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.value = ""
        cell.border = border

# Função para adicionar os dados de horas na planilha
def adicionar_dados(linha, data, hora_ini, hora_fin, multiplicador, delta_horas, motivo):
    ws.cell(row=linha, column=1, value=data).border = border
    ws.cell(row=linha, column=2, value=decimal_para_hora(hora_ini)).border = border
    ws.cell(row=linha, column=3, value=decimal_para_hora(hora_fin)).border = border
    ws.cell(row=linha, column=4, value=multiplicador).border = border
    ws.cell(row=linha, column=5, value=delta_horas).border = border
    ws.cell(row=linha, column=6, value=motivo).border = border  # Adicionando o motivo na coluna F

# Loop para preencher as horas extras
linha_atual = 5
total_horas = 0

while linha_atual <= 20:
    data, hora_ini, hora_fin, multiplicador, delta_horas, motivo = obter_dados_horas(linha_atual)
    adicionar_dados(linha_atual, data, hora_ini, hora_fin, multiplicador, delta_horas, motivo)
    
    # Atualizando o total de horas
    total_horas += hora_fin - hora_ini
    
    # Perguntar se o usuário tem mais dados
    mais_dados = input("Você tem mais horas extras para adicionar? (s/n): ").strip().lower()
    if mais_dados != 's':
        break
    
    # Avançando para a próxima linha
    linha_atual += 1

# Colocando o total de horas em 21E
ws.cell(row=21, column=5, value=decimal_para_hora(total_horas)).border = border
set_uppercase(ws['D21'], "TOTAL")
set_uppercase(ws['E21'], decimal_para_hora(total_horas))

# Ajustando a largura das colunas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 65

# Aplicando o estilo de fonte (negrito e tamanho 14) nas células preenchidas
for row in range(5, linha_atual+1):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = bold_font  # Definindo a fonte em negrito e tamanho 14
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Centralizando o texto
        cell.border = border

# Formatação de horário para a coluna B, C e 21E
for row in range(5, linha_atual+1):
    for col in [2, 3]:  # Colunas B e C (hora inicial e hora final)
        cell = ws.cell(row=row, column=col)
        cell.number_format = 'HH:MM:SS'  # Formatando as células como hora

# Salvar o arquivo
file_name = "hora_extra.xlsx"
wb.save(file_name)
print(f"Planilha salva como {file_name}.")

# 1. Enviar o arquivo por e-mail
enviarEmail(file_name)

# 2. Fazer upload para o Dropbox e obter o link
dropbox_link = enviar_para_dropbox(file_name)
print(f"Arquivo enviado para o Dropbox. Link compartilhável: {dropbox_link}")

# 3. Enviar o link para o WhatsApp
numero_destinatario = "+558588570694"  # Substitua pelo número correto
mensagem = f"Segue o link para o arquivo de horas extras: {dropbox_link}"

try:
    kit.sendwhatmsg_instantly(numero_destinatario, mensagem, wait_time=15)  # Define horário para envio
    print("Mensagem enviada com sucesso pelo WhatsApp!")
except Exception as e:
    print(f"Erro ao enviar mensagem no WhatsApp: {e}")